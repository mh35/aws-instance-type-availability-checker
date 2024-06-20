"""まとまった情報をExcel形式にして、S3バケットに保存します."""

import os
import tempfile
from typing import TYPE_CHECKING, TypedDict

import boto3
import openpyxl

if TYPE_CHECKING:
    from aws_lambda_powertools.utilities.typing import LambdaContext

class EventDict(TypedDict):
    """イベントの一部の辞書です."""

    region: str
    offerings: dict[str, list[str]]

def handler(event: list[EventDict], context: "LambdaContext") -> dict[str, str]:
    """出力を生成します."""
    sess = boto3.Session()
    s3 = sess.resource("s3")
    bucket = s3.Bucket(os.environ["BUCKET_NAME"])
    workbook = openpyxl.Workbook()
    is_first = True
    for region in event:
        if is_first:
            is_first = False
            worksheet = workbook.worksheets[0]
            worksheet.title = region["region"]
        else:
            worksheet = workbook.create_sheet(region["region"])
        azs: set[str] = set()
        for offering_azs in region["offerings"].values():
            azs |= set(offering_azs)
        az_list = sorted(list(azs))
        worksheet.cell(2, 2).value = "Type"
        for ci, az in enumerate(az_list):
            worksheet.cell(2, ci + 3).value = az
        cidx = 3
        for itype, oazs in region["offerings"].items():
            worksheet.cell(cidx, 2).value = itype
            for ci, taz in enumerate(az_list):
                worksheet.cell(cidx, ci + 3).value = 'o' if taz in oazs else 'x'
            cidx += 1
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as fp:
        workbook.save(fp.name)
        bucket.put_object(
            Key="instance_type_offerings.xlsx",
            Body=fp,
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    return {"status": "success"}
